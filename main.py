import openpyxl

# Load the workbook and select the sheet
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Dictionaries to hold results
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# Print the number of rows
print(product_list.max_row)

# Add a header for the new column
product_list.cell(row=1, column=5).value = "Total Inventory Value"

# Iterate through rows, starting from the second row
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(row=product_row, column=4).value
    inventory = product_list.cell(row=product_row, column=2).value
    price = product_list.cell(row=product_row, column=3).value
    products_num = product_list.cell(row=product_row, column=1).value
    inventory_price_cell = product_list.cell(row=product_row, column=5)

    # Calculation number of products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1

    # Calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(products_num)] = int(inventory)

    # Add value for total inventory price
    inventory_price_cell.value = inventory * price

# Print the dictionaries
print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

# Save the workbook with a new name
inv_file.save("inventory_with_total_value.xlsx")
